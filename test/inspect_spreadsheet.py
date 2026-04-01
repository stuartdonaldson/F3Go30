#!/usr/bin/env python3
"""
Extract the full structure and content of an F3Go30 spreadsheet.

Discovers the spreadsheet from the most recent successful copyAndInit log entry,
downloads it as XLSX, and extracts all sheets with cell values, formulas, and
key formatting attributes.

Usage:
    # Discover from logfile (same approach as test_tracker_init.py)
    python inspect_spreadsheet.py --log <drive_logfile_url>

    # Direct spreadsheet URL
    python inspect_spreadsheet.py <spreadsheet_url>

Options:
    --out <file>   Write JSON output to file (default: stdout)
    --doc <file>   Write Markdown sheet-reference document to file

Output JSON structure:
    {
      "spreadsheet_url": "...",
      "log_payload": {...},          # null when not using --log
      "sheets": [
        {
          "name": "...",
          "state": "visible|hidden|veryHidden",
          "tab_color": "RRGGBB or null",
          "freeze_row": N,           # rows frozen (0 = none)
          "freeze_col": N,           # cols frozen (0 = none)
          "max_row": N,
          "max_col": N,
          "merged_cells": ["A1:B2", ...],
          "col_dims": {"A": {"width": 10.0, "hidden": false}, ...},
          "row_dims": {"1": {"height": 15.0, "hidden": false}, ...},
          "cells": [
            {
              "row": 1, "col": 1, "address": "A1",
              "value": "...",        # null when cell contains a formula
              "formula": "=...",     # null when cell contains a plain value
              "bold": true,
              "italic": false,
              "bg_color": "RRGGBB or null",
              "font_color": "RRGGBB or null",
              "align_h": "center",
              "align_v": "bottom",
              "wrap": false,
              "number_format": "General"
            },
            ...
          ]
        },
        ...
      ]
    }
"""

import io
import json
import re
import sys
from typing import Any

import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.formula import ArrayFormula

sys.path.insert(0, __file__.rsplit("/", 1)[0])
from log_channel import fetch_log_entries


# ---------------------------------------------------------------------------
# Shared helpers (same pattern as test_tracker_init.py)
# ---------------------------------------------------------------------------

def extract_sheet_id(url: str) -> str:
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    if not m:
        raise ValueError(f"Cannot extract sheet ID from URL: {url}")
    return m.group(1)


def build_export_url(sheet_id: str) -> str:
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"


def download_xlsx(export_url: str) -> bytes:
    resp = requests.get(export_url, timeout=30)
    resp.raise_for_status()
    return resp.content


def follow_redirect(url: str) -> str:
    resp = requests.head(url, allow_redirects=True, timeout=15)
    return resp.url


def find_sheet_url_from_log(log_url: str) -> tuple[str, dict]:
    """
    Find the most recent successful copyAndInit log entry and return
    (spreadsheet_url, payload).
    """
    entries = fetch_log_entries(log_url)
    if not entries:
        raise RuntimeError("LogFile is empty — no entries found.")

    for entry in reversed(entries):
        if entry["trigger"] != "copyAndInit":
            continue
        payload = entry["payload"]
        if "error" in payload:
            continue
        tracker_url = payload.get("trackerUrl")
        if not tracker_url:
            continue

        print(f"Most recent copyAndInit: [{entry['timestamp']}]", file=sys.stderr)
        print(f"  spreadsheetName: {payload.get('spreadsheetName', '?')}", file=sys.stderr)
        print(f"  trackerUrl: {tracker_url}", file=sys.stderr)

        final_url = follow_redirect(tracker_url)
        print(f"  resolved to: {final_url}", file=sys.stderr)
        sheet_url = final_url.split("#")[0]
        return sheet_url, payload

    raise RuntimeError("No successful copyAndInit entry found in LogFile.")


# ---------------------------------------------------------------------------
# Color helpers
# ---------------------------------------------------------------------------

def _rgb_str(color_obj) -> str | None:
    """Return RRGGBB hex string from an openpyxl Color, or None."""
    if color_obj is None:
        return None
    try:
        if color_obj.type == "rgb":
            rgb = color_obj.rgb.upper()  # AARRGGBB
            if len(rgb) == 8:
                return rgb[2:]  # strip alpha → RRGGBB
            return rgb
    except Exception:
        pass
    return None


# ---------------------------------------------------------------------------
# Extraction
# ---------------------------------------------------------------------------

def extract_cell(cell) -> dict[str, Any] | None:
    """Return a dict for a cell, or None if empty."""
    value = cell.value
    if value is None:
        return None

    # ArrayFormula objects wrap array-entered formulas; extract the formula text
    if isinstance(value, ArrayFormula):
        value = value.text or str(value)

    formula = None
    plain_value = value
    if isinstance(value, str) and value.startswith("="):
        formula = value
        plain_value = None

    bold = False
    italic = False
    bg_color = None
    font_color = None
    align_h = None
    align_v = None
    wrap = False
    number_format = cell.number_format or "General"

    if cell.font:
        bold = bool(cell.font.bold)
        italic = bool(cell.font.italic)
        if cell.font.color:
            font_color = _rgb_str(cell.font.color)

    if cell.fill and cell.fill.patternType not in (None, "none"):
        if cell.fill.fgColor:
            bg_color = _rgb_str(cell.fill.fgColor)

    if cell.alignment:
        align_h = cell.alignment.horizontal
        align_v = cell.alignment.vertical
        wrap = bool(cell.alignment.wrapText)

    return {
        "row": cell.row,
        "col": cell.column,
        "address": cell.coordinate,
        "value": plain_value,
        "formula": formula,
        "bold": bold,
        "italic": italic,
        "bg_color": bg_color,
        "font_color": font_color,
        "align_h": align_h,
        "align_v": align_v,
        "wrap": wrap,
        "number_format": number_format,
    }


def extract_sheet(ws) -> dict[str, Any]:
    """Extract full structure and content from a worksheet."""
    # Freeze panes — e.g. "B4" means cols < B and rows < 4 are frozen
    freeze_row = 0
    freeze_col = 0
    if ws.freeze_panes:
        m = re.fullmatch(r"([A-Z]+)(\d+)", str(ws.freeze_panes))
        if m:
            freeze_col = column_index_from_string(m.group(1)) - 1
            freeze_row = int(m.group(2)) - 1

    # Tab color
    tab_color = None
    if ws.sheet_properties and ws.sheet_properties.tabColor:
        tab_color = _rgb_str(ws.sheet_properties.tabColor)

    # Merged cell regions
    merged = [str(r) for r in ws.merged_cells.ranges]

    # Column dimensions
    col_dims: dict[str, Any] = {}
    for col_letter, dim in ws.column_dimensions.items():
        col_dims[col_letter] = {
            "width": dim.width,
            "hidden": bool(dim.hidden),
        }

    # Row dimensions
    row_dims: dict[str, Any] = {}
    for row_num, dim in ws.row_dimensions.items():
        row_dims[str(row_num)] = {
            "height": dim.height,
            "hidden": bool(dim.hidden),
        }

    # Cells (skip entirely empty cells)
    cells = []
    for row in ws.iter_rows():
        for cell in row:
            c = extract_cell(cell)
            if c is not None:
                cells.append(c)

    return {
        "name": ws.title,
        "state": ws.sheet_state,
        "tab_color": tab_color,
        "freeze_row": freeze_row,
        "freeze_col": freeze_col,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "merged_cells": merged,
        "col_dims": col_dims,
        "row_dims": row_dims,
        "cells": cells,
    }


def extract_workbook(xlsx_bytes: bytes) -> list[dict[str, Any]]:
    # data_only=False preserves formula strings instead of cached values
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=False)
    sheets = []
    for name in wb.sheetnames:
        print(f"  Extracting: {name}", file=sys.stderr)
        sheets.append(extract_sheet(wb[name]))
    return sheets


# ---------------------------------------------------------------------------
# Documentation generation
# ---------------------------------------------------------------------------

def _cell_map(sheet: dict) -> dict[str, str]:
    """Address → display string (formula or value)."""
    result = {}
    for c in sheet["cells"]:
        display = c.get("formula") or c.get("value")
        if display is not None:
            result[c["address"]] = str(display)
    return result


def generate_doc(sheets: list[dict], spreadsheet_url: str) -> str:
    lines = [
        "# F3Go30 Spreadsheet Reference",
        "",
        f"Source: {spreadsheet_url}",
        "",
        "---",
        "",
    ]

    for sheet in sheets:
        cm = _cell_map(sheet)
        name = sheet["name"]
        state = sheet["state"]
        max_row = sheet["max_row"] or 0
        max_col = sheet["max_col"] or 0

        lines.append(f"## Sheet: {name}")
        lines.append("")
        lines.append(f"- **Visibility:** {state}")
        lines.append(f"- **Dimensions:** {max_row} rows × {max_col} columns")

        if sheet["freeze_row"] or sheet["freeze_col"]:
            lines.append(
                f"- **Frozen panes:** {sheet['freeze_row']} row(s), "
                f"{sheet['freeze_col']} column(s)"
            )

        if sheet["tab_color"]:
            lines.append(f"- **Tab color:** #{sheet['tab_color']}")

        merged = sheet["merged_cells"]
        if merged:
            lines.append(f"- **Merged regions ({len(merged)}):** {', '.join(merged[:10])}"
                         + (" …" if len(merged) > 10 else ""))

        # Row 1 — headers/labels
        row1 = [
            f"{get_column_letter(col)}: {cm[f'{get_column_letter(col)}1']}"
            for col in range(1, max_col + 1)
            if f"{get_column_letter(col)}1" in cm
        ]
        if row1:
            lines.append("")
            lines.append("### Row 1")
            lines.append("")
            for h in row1:
                lines.append(f"- {h}")

        # Row 2 — sub-headers if different
        row2 = [
            f"{get_column_letter(col)}: {cm[f'{get_column_letter(col)}2']}"
            for col in range(1, max_col + 1)
            if f"{get_column_letter(col)}2" in cm
        ]
        if row2 and row2 != row1:
            lines.append("")
            lines.append("### Row 2")
            lines.append("")
            for h in row2:
                lines.append(f"- {h}")

        # Row 3 or 4 — sample/template row
        for sample_row in (3, 4):
            sample = [
                f"{get_column_letter(col)}{sample_row}: "
                f"{cm[f'{get_column_letter(col)}{sample_row}']}"
                for col in range(1, min(max_col + 1, 25))
                if f"{get_column_letter(col)}{sample_row}" in cm
            ]
            if sample:
                lines.append("")
                lines.append(f"### Row {sample_row} (sample / template)")
                lines.append("")
                for s in sample:
                    lines.append(f"- {s}")
                break

        # Unique formulas
        formulas_seen: set[str] = set()
        formulas_list: list[str] = []
        for c in sheet["cells"]:
            f = c.get("formula")
            if f and f not in formulas_seen:
                formulas_seen.add(f)
                formulas_list.append(f"{c['address']}: {f}")

        if formulas_list:
            lines.append("")
            lines.append(f"### Formulas ({len(formulas_list)} unique)")
            lines.append("")
            lines.append("```")
            for f in formulas_list[:60]:
                lines.append(f)
            if len(formulas_list) > 60:
                lines.append(f"... and {len(formulas_list) - 60} more")
            lines.append("```")

        # Hidden columns
        hidden_cols = sorted(
            col for col, dim in sheet["col_dims"].items() if dim.get("hidden")
        )
        if hidden_cols:
            lines.append("")
            lines.append(f"### Hidden columns: {', '.join(hidden_cols)}")

        # Hidden rows
        hidden_rows = sorted(
            int(r) for r, dim in sheet["row_dims"].items() if dim.get("hidden")
        )
        if hidden_rows:
            lines.append("")
            lines.append(f"### Hidden rows: {', '.join(str(r) for r in hidden_rows)}")

        lines.append("")
        lines.append("---")
        lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    args = sys.argv[1:]
    log_url = None
    sheet_url = None
    out_file = None
    doc_file = None

    i = 0
    while i < len(args):
        if args[i] == "--log" and i + 1 < len(args):
            log_url = args[i + 1]
            i += 2
        elif args[i] == "--out" and i + 1 < len(args):
            out_file = args[i + 1]
            i += 2
        elif args[i] == "--doc" and i + 1 < len(args):
            doc_file = args[i + 1]
            i += 2
        elif not args[i].startswith("--"):
            sheet_url = args[i]
            i += 1
        else:
            print(f"Unknown argument: {args[i]}", file=sys.stderr)
            sys.exit(2)

    if log_url is None and sheet_url is None:
        print("Usage:", file=sys.stderr)
        print(
            f"  {sys.argv[0]} --log <drive_logfile_url>"
            " [--out data.json] [--doc sheet-reference.md]",
            file=sys.stderr,
        )
        print(
            f"  {sys.argv[0]} <spreadsheet_url>"
            " [--out data.json] [--doc sheet-reference.md]",
            file=sys.stderr,
        )
        sys.exit(2)

    log_payload = None
    if log_url:
        print(f"LogFile URL: {log_url}", file=sys.stderr)
        try:
            sheet_url, log_payload = find_sheet_url_from_log(log_url)
        except RuntimeError as e:
            print(f"ERROR: {e}", file=sys.stderr)
            sys.exit(2)

    print(f"Sheet URL: {sheet_url}", file=sys.stderr)

    try:
        sheet_id = extract_sheet_id(sheet_url)
    except ValueError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(2)

    export_url = build_export_url(sheet_id)
    print(f"Downloading: {export_url}", file=sys.stderr)

    try:
        xlsx_bytes = download_xlsx(export_url)
    except requests.HTTPError as e:
        print(f"ERROR downloading spreadsheet: {e}", file=sys.stderr)
        sys.exit(2)

    print("Extracting sheets...", file=sys.stderr)
    sheets = extract_workbook(xlsx_bytes)

    result: dict[str, Any] = {
        "spreadsheet_url": sheet_url,
        "log_payload": log_payload,
        "sheets": sheets,
    }

    json_text = json.dumps(result, indent=2, default=str)

    if out_file:
        with open(out_file, "w", encoding="utf-8") as fh:
            fh.write(json_text)
        print(f"JSON written to: {out_file}", file=sys.stderr)
    else:
        print(json_text)

    if doc_file:
        doc_text = generate_doc(sheets, sheet_url)
        with open(doc_file, "w", encoding="utf-8") as fh:
            fh.write(doc_text)
        print(f"Documentation written to: {doc_file}", file=sys.stderr)


if __name__ == "__main__":
    main()

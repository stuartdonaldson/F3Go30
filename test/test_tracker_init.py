#!/usr/bin/env python3
"""
Verify that a Google Sheets Tracker has been correctly initialized by copyAndInit().

Usage:
    # Primary: discover sheet URL from the most recent local GasLogger entry
    python test_tracker_init.py --local

    # Legacy: discover sheet URL from the most recent log entry in the old LogFile
    python test_tracker_init.py --log <drive_logfile_url>

    # Legacy: provide sheet URL directly
    python test_tracker_init.py <google_sheets_url>

The sheet must be publicly shared (no auth required).
Exits 0 if all checks pass, non-zero if any fail.
"""

import sys
import re
import io
import calendar
from datetime import date, datetime as dt, timedelta

import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Import log_channel from same directory
sys.path.insert(0, __file__.rsplit("/", 1)[0])
from log_channel import _load_settings, collect_local_log_entries, fetch_log_entries


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def extract_sheet_id(url: str) -> str:
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    if not m:
        raise ValueError(f"Cannot extract sheet ID from URL: {url}")
    return m.group(1)


def build_export_url(sheet_id: str) -> str:
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"


def download_xlsx(export_url: str) -> bytes:
    resp = requests.get(export_url, timeout=30)
    resp.raise_for_status()
    return resp.content


def follow_redirect(url: str) -> str:
    """Follow redirects and return the final URL."""
    resp = requests.head(url, allow_redirects=True, timeout=15)
    return resp.url


def stringify_cell(value) -> str:
    if isinstance(value, dt):
        value = value.date()
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value or "").strip()


def build_header_map(ws) -> dict[str, int]:
    headers = {}
    for col in range(1, ws.max_column + 1):
        header = stringify_cell(ws.cell(row=1, column=col).value).lower()
        if header:
            headers[header] = col
    return headers


def read_sheet_rows(ws, header_map: dict[str, int]) -> list[dict[str, str]]:
    rows = []
    for row in range(2, ws.max_row + 1):
        row_data = {
            header: stringify_cell(ws.cell(row=row, column=col).value)
            for header, col in header_map.items()
        }
        if any(value for value in row_data.values()):
            rows.append(row_data)
    return rows


def find_previous_tracker_from_links(rows: list[dict[str, str]], start_date_iso: str):
    start_date = date.fromisoformat(start_date_iso)
    previous_month = date(start_date.year, start_date.month, 1) - timedelta(days=1)
    previous_month_key = previous_month.strftime("%Y-%m")

    for row in reversed(rows):
        row_start_date = row.get("startdate", "").strip()
        if row_start_date[:7] != previous_month_key:
            continue
        return {
            "name": row_start_date,
            "url": row.get("trackerurl", "").strip(),
        }

    return None


def find_sheet_url_from_log(log_url: str) -> tuple[str, dict]:
    """
    Download the LogFile, find the most recent successful copyAndInit entry,
    follow the trackerUrl redirect to obtain the full spreadsheet URL.

    Returns (sheet_url, entry) where entry is the log payload dict.
    """
    entries = fetch_log_entries(log_url)
    if not entries:
        raise RuntimeError("LogFile is empty — no entries found.")

    # Most recent entry is last; search in reverse for a successful copyAndInit
    for entry in reversed(entries):
        if entry["trigger"] != "copyAndInit":
            continue
        payload = entry["payload"]
        if "error" in payload:
            continue
        tracker_url = payload.get("trackerUrl")
        if not tracker_url:
            continue

        print(f"Most recent copyAndInit: [{entry['timestamp']}]")
        print(f"  spreadsheetName: {payload.get('spreadsheetName', '?')}")
        print(f"  trackerUrl: {tracker_url}")

        # Follow the short URL redirect to get the full spreadsheet URL
        final_url = follow_redirect(tracker_url)
        print(f"  resolved to: {final_url}")

        # Strip the #gid fragment — we need the spreadsheet root URL
        sheet_url = final_url.split("#")[0]
        return sheet_url, payload

    raise RuntimeError(
        "No successful copyAndInit entry found in LogFile. "
        "Most recent entries:\n" +
        "\n".join(
            f"  [{e['timestamp']}] {e['trigger']} "
            f"{'(error)' if 'error' in e['payload'] else ''}"
            for e in reversed(entries[-5:])
        )
    )


def find_sheet_url_from_local_logs() -> tuple[str, dict]:
    settings = _load_settings()
    local_path = settings.get("GAS_LOGGER_LOCAL_PATH")
    prefix = settings.get("GAS_LOGGER_PROJECT_PREFIX", "F3Go30")
    if not local_path:
        raise RuntimeError("local.settings.json is missing GAS_LOGGER_LOCAL_PATH.")

    scan_path = str((__import__("pathlib").Path(local_path) / prefix))
    files_map = collect_local_log_entries(scan_path)
    if not files_map:
        raise RuntimeError(f"No local log entries found under {scan_path}.")

    candidates = []
    for file_path, entries in files_map.items():
        for entry in entries:
            if entry.get("tag") != "copyAndInit":
                continue
            payload = entry.get("data") or {}
            if payload.get("error"):
                continue
            tracker_url = payload.get("trackerUrl")
            if not tracker_url:
                continue
            candidates.append((entry.get("ts") or "", file_path, payload))

    if not candidates:
        raise RuntimeError("No successful copyAndInit entries found in local GasLogger files.")

    timestamp, file_path, payload = max(candidates, key=lambda item: item[0])
    print(f"Most recent local copyAndInit: [{timestamp}]")
    print(f"  log file: {file_path}")
    print(f"  spreadsheetName: {payload.get('spreadsheetName', '?')}")
    print(f"  trackerUrl: {payload.get('trackerUrl')}")

    final_url = follow_redirect(payload["trackerUrl"])
    print(f"  resolved to: {final_url}")
    return final_url.split("#")[0], payload


def argb_matches(fill, hex_color: str) -> bool:
    """Check if an openpyxl PatternFill matches a hex color (RRGGBB or FFRRGGBB)."""
    if fill is None or fill.fgColor is None:
        return False
    color = fill.fgColor
    if color.type == "rgb":
        rgb = color.rgb.upper()
        if len(rgb) == 8:
            rgb = rgb[2:]  # strip alpha byte (first two hex chars)
        return rgb == hex_color.upper().lstrip("#")
    return False


ORANGE = "FF9900"
GREEN  = "00FF00"
LAST_TRACKER_COLUMN = 44  # Column AR — matches GAS LAST_TRACKER_COLUMN constant

RESULT = {"pass": 0, "fail": 0}


def check(name: str, passed: bool, detail: str = "") -> bool:
    status = "PASS" if passed else "FAIL"
    msg = f"  [{status}] {name}"
    if detail:
        msg += f": {detail}"
    print(msg)
    RESULT["pass" if passed else "fail"] += 1
    return passed


# ---------------------------------------------------------------------------
# Determine expected month from the first date in row 3
# ---------------------------------------------------------------------------

def first_date_in_row3(ws):
    """Return the first date value found in row 3, or None (always a date, never datetime)."""
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=3, column=col).value
        if isinstance(val, dt):
            return val.date()
        if isinstance(val, date):
            return val
        if isinstance(val, str):
            try:
                parts = val.split("/")
                if len(parts) == 2:
                    return date(date.today().year, int(parts[0]), int(parts[1]))
            except (ValueError, IndexError):
                pass
    return None


# ---------------------------------------------------------------------------
# Checks
# ---------------------------------------------------------------------------

def check_tracker(ws):
    print("\n--- Tracker sheet ---")

    # Identify the first date cell to determine the month being tested
    first_date = first_date_in_row3(ws)
    if not check("Row 3 has at least one date value", first_date is not None):
        return

    year = first_date.year
    month = first_date.month
    days_in_month = calendar.monthrange(year, month)[1]
    month_start = date(year, month, 1)

    # Find the column where the first DATE sits (skip fixed header columns like "F3 Name")
    first_date_col = None
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=3, column=col).value
        if isinstance(val, date):
            first_date_col = col
            break
        if isinstance(val, str):
            try:
                parts = val.split("/")
                if len(parts) == 2:
                    int(parts[0]); int(parts[1])
                    first_date_col = col
                    break
            except (ValueError, IndexError):
                pass

    if first_date_col is None:
        check("Could anchor date column in row 3", False)
        return

    # AC 4: sequential dates starting from the 1st
    date_cols = {}        # date → column index
    bonus_cols = set()    # column indices that are Bonus columns
    bad_dates = []

    expected_day = 1
    col = first_date_col
    while expected_day <= days_in_month and col <= ws.max_column:
        cell = ws.cell(row=3, column=col)
        val = cell.value

        # Detect a Bonus column (header says "Bonus" or similar, or green fill)
        header = ws.cell(row=1, column=col).value or ws.cell(row=2, column=col).value or ""
        fill = cell.fill
        is_bonus = (
            (isinstance(header, str) and "bonus" in header.lower())
            or argb_matches(fill, GREEN)
        )
        if is_bonus:
            bonus_cols.add(col)
            col += 1
            continue

        # Parse the date value (normalize datetime → date)
        if isinstance(val, dt):
            cell_date = val.date()
        elif isinstance(val, date):
            cell_date = val
        elif isinstance(val, str):
            try:
                parts = val.split("/")
                cell_date = date(year, int(parts[0]), int(parts[1]))
            except (ValueError, IndexError, TypeError):
                bad_dates.append((col, val))
                col += 1
                expected_day += 1
                continue
        else:
            bad_dates.append((col, val))
            col += 1
            expected_day += 1
            continue

        expected_date = month_start + timedelta(days=expected_day - 1)
        if cell_date != expected_date:
            bad_dates.append((col, f"{val!r} expected {expected_date.strftime('%m/%d')}"))
        else:
            date_cols[cell_date] = col

        expected_day += 1
        col += 1

    check("Row 3 contains sequential dates for the month",
          len(bad_dates) == 0,
          f"mismatches: {bad_dates}" if bad_dates else "")

    # AC 4 format: MM/DD
    bad_fmt = []
    for col_idx in date_cols.values():
        val = ws.cell(row=3, column=col_idx).value
        if isinstance(val, str):
            if not re.fullmatch(r"\d{1,2}/\d{1,2}", val):
                bad_fmt.append((col_idx, val))
        # date objects are formatted by Sheets; xlsx may render as date type — acceptable
    check("Date cells formatted MM/DD (or date type)",
          len(bad_fmt) == 0,
          f"bad format: {bad_fmt}" if bad_fmt else "")

    # AC 5: orange fill on date columns
    non_orange = []
    for d, col_idx in date_cols.items():
        fill = ws.cell(row=3, column=col_idx).fill
        if not argb_matches(fill, ORANGE):
            non_orange.append(col_idx)
    check("Date cells have orange fill (#FF9900)",
          len(non_orange) == 0,
          f"cols without orange: {non_orange}" if non_orange else "")

    # AC 6: each Saturday is followed by a Bonus column with green fill
    bad_bonus = []
    for d, col_idx in date_cols.items():
        if d.weekday() == 5:  # Saturday
            next_col = col_idx + 1
            bonus_fill = ws.cell(row=3, column=next_col).fill
            if not argb_matches(bonus_fill, GREEN):
                bad_bonus.append(f"col {next_col} (after Sat {d.strftime('%m/%d')})")
    check("Each Saturday is followed by a Bonus column (green fill)",
          len(bad_bonus) == 0,
          f"missing: {bad_bonus}" if bad_bonus else "")

    # AC 7 (trailing Bonus): a Bonus column exists after the last day of the month
    if date_cols:
        last_date = max(date_cols.keys())
        last_col = date_cols[last_date]
        trailing_col = last_col + 1
        trailing_fill = ws.cell(row=3, column=trailing_col).fill
        trailing_header = (
            ws.cell(row=1, column=trailing_col).value or
            ws.cell(row=2, column=trailing_col).value or ""
        )
        has_trailing = (
            argb_matches(trailing_fill, GREEN)
            or (isinstance(trailing_header, str) and "bonus" in trailing_header.lower())
        )
        check("Trailing Bonus column exists after last day of month",
              has_trailing,
              f"col {trailing_col} fill={ws.cell(row=3, column=trailing_col).fill.fgColor.rgb!r}"
              if not has_trailing else "")

    # AC 8: no data rows beyond row 4
    non_empty = []
    for row in range(5, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            if ws.cell(row=row, column=col_idx).value not in (None, ""):
                non_empty.append(row)
                break
    check("No data rows beyond row 4 in Tracker",
          len(non_empty) == 0,
          f"rows with data: {non_empty}" if non_empty else "")

    # AC 9 (partial — hidden columns): within the dynamic tracker window I..AS,
    # the first unused column after the date/bonus region must be hidden.
    if date_cols:
        last_date_col = max(date_cols.values())
        check_from = last_date_col + 1
        while check_from in bonus_cols or argb_matches(ws.cell(row=3, column=check_from).fill, GREEN):
            check_from += 1
        last_dynamic_column = 45  # AS
        if check_from > last_dynamic_column:
            check("First column after date area is hidden", True,
                  "month fills the dynamic tracker window")
        else:
            col_letter = get_column_letter(check_from)
            dim = ws.column_dimensions.get(col_letter)
            is_hidden = (dim.hidden if dim else False)
            check("First column after date area is hidden",
                  is_hidden,
                  f"col {check_from} ({col_letter}) not hidden" if not is_hidden else "")


def check_bonus_tracker(wb):
    print("\n--- Bonus Tracker sheet ---")
    if "Bonus Tracker" not in wb.sheetnames:
        check("Bonus Tracker sheet exists", False, "sheet not found")
        return
    ws = wb["Bonus Tracker"]
    non_empty = []
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=row, column=col).value not in (None, ""):
                non_empty.append(row)
                break
    check("Bonus Tracker row 2+ is empty",
          len(non_empty) == 0,
          f"rows with data: {non_empty}" if non_empty else "")


def check_responses(wb):
    print("\n--- Responses sheet ---")
    sheet_name = next((s for s in wb.sheetnames if "response" in s.lower()), None)
    if sheet_name is None:
        check("Responses sheet exists", False, "sheet not found")
        return
    ws = wb[sheet_name]
    non_empty = []
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=row, column=col).value not in (None, ""):
                non_empty.append(row)
                break
    check("Responses sheet has only the header row",
          len(non_empty) == 0,
          f"rows with data: {non_empty}" if non_empty else "")


def check_activity(wb):
    print("\n--- Activity sheet ---")
    sheet_name = next((s for s in wb.sheetnames if "activity" in s.lower()), None)
    if sheet_name is None:
        check("Activity sheet exists", False, "sheet not found")
        return
    ws = wb[sheet_name]
    unexpected_rows = []
    for row in range(2, ws.max_row + 1):
        values = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
        if all(value in (None, "") for value in values):
            continue
        message = stringify_cell(values[2]) if len(values) >= 3 else ""
        if row == 2 and message == "onOpen":
            continue
        unexpected_rows.append((row, values))
    check("Activity sheet has no unexpected residual rows",
          len(unexpected_rows) == 0,
          f"rows with data: {unexpected_rows}" if unexpected_rows else "")


def check_log_payload(payload: dict):
    """Assert LogFile payload contains the lineage fields needed for live verification."""
    print("\n--- LogFile payload ---")
    required_fields = [
      "spreadsheetId",
      "spreadsheetName",
      "startDateIso",
      "trackerUrl",
      "formUrl",
      "templateSpreadsheetId",
    ]
    for field in required_fields:
      check(f"LogFile payload contains {field}",
          payload.get(field) not in (None, ""),
          "key missing from payload" if field not in payload else f"value={payload.get(field)!r}")


def check_links_sheet(payload: dict):
    """Download the template spreadsheet and verify the Links sheet row."""
    print("\n--- Links sheet (template) ---")
    template_id = payload.get("templateSpreadsheetId")
    if not check("templateSpreadsheetId present in payload", template_id is not None,
                 "key missing — cannot verify Links sheet"):
        return

    export_url = build_export_url(template_id)
    print(f"  Downloading template: {export_url}")
    try:
        xlsx_bytes = download_xlsx(export_url)
    except requests.HTTPError as e:
        check("Template spreadsheet downloadable", False, str(e))
        return

    wb = load_workbook(io.BytesIO(xlsx_bytes))
    if not check("Links sheet exists in template", "Links" in wb.sheetnames):
        return None

    ws = wb["Links"]
    header_map = build_header_map(ws)
    required_headers = ["date", "startdate", "shorttracker", "trackerurl", "shorthc", "hc url", "sheetid"]
    missing_headers = [header for header in required_headers if header not in header_map]
    if not check("Links sheet has lineage headers", len(missing_headers) == 0,
                 f"missing headers: {missing_headers}" if missing_headers else ""):
        return None

    rows = read_sheet_rows(ws, header_map)
    target_sheet_id = str(payload.get("spreadsheetId") or "").strip()
    target_month = str(payload.get("startDateIso") or "").strip()
    target_short_tracker = str(payload.get("trackerUrl") or "").strip()
    target_short_form = str(payload.get("formUrl") or "").strip()
    resolved_tracker = follow_redirect(target_short_tracker) if target_short_tracker else ""
    resolved_form = follow_redirect(target_short_form) if target_short_form else ""

    matching_rows = [row for row in rows if row.get("sheetid", "").strip() == target_sheet_id]
    if not check("Links sheet has exactly one authoritative row for this tracker",
                 len(matching_rows) == 1,
                 f"rows with sheetId={target_sheet_id!r}: {len(matching_rows)}"):
        return None

    found_row = matching_rows[0]
    row_date = found_row.get("date", "")
    row_month = found_row.get("startdate", "")
    row_short_tracker = found_row.get("shorttracker", "")
    row_tracker = found_row.get("trackerurl", "")
    row_short_form = found_row.get("shorthc", "")
    row_form = found_row.get("hc url", "")
    row_sheet_id = found_row.get("sheetid", "")

    check("Links row has non-empty date", row_date not in (None, ""), f"date={row_date!r}")
    check("Links row month matches startDateIso", row_month == target_month,
          f"got={row_month!r} expected={target_month!r}")
    check("Links row sheetId matches payload", row_sheet_id == target_sheet_id,
          f"got={row_sheet_id!r} expected={target_sheet_id!r}")
    check("Links row shortTracker matches payload", row_short_tracker == target_short_tracker,
          f"got={row_short_tracker!r} expected={target_short_tracker!r}")
    check("Links row trackerUrl matches resolved short URL", row_tracker == resolved_tracker,
          f"got={row_tracker!r} expected={resolved_tracker!r}")
    check("Links row short HC URL matches payload", row_short_form == target_short_form,
          f"got={row_short_form!r} expected={target_short_form!r}")
    check("Links row HC URL matches resolved short URL", row_form == resolved_form,
          f"got={row_form!r} expected={resolved_form!r}")

    return {
        "template_sheet_id": str(template_id).strip(),
        "previous_tracker": find_previous_tracker_from_links(rows, target_month),
    }


def check_config(wb, payload: dict | None = None, template_lineage: dict | None = None):
    print("\n--- Config sheet ---")
    sheet_name = next((s for s in wb.sheetnames if s.lower() == "config"), None)
    if sheet_name is None:
        check("Config sheet exists", False, "sheet not found")
        return
    ws = wb[sheet_name]

    # AC 12: Config sheet should be hidden
    state = ws.sheet_state  # 'visible', 'hidden', 'veryHidden'
    check("Config sheet is hidden",
          state in ("hidden", "veryHidden"),
          f"sheet_state={state!r}")

    # Build a map of variable name → (col_b, col_c)
    config_rows = {}
    for row in range(1, ws.max_row + 1):
        key = ws.cell(row=row, column=1).value
        if key:
            config_rows[str(key).strip()] = (
                ws.cell(row=row, column=2).value,
                ws.cell(row=row, column=3).value,
            )

    # AC 13: NameSpace row with non-empty value in column B
    ns = config_rows.get("NameSpace")
    check("Config has NameSpace row with non-empty column B",
          ns is not None and ns[0] not in (None, ""),
          f"value={ns!r}" if ns is not None else "row not found")

    # AC 13: Site Q row with name in column B and email in column C
    sq = config_rows.get("Site Q")
    has_sq_name = sq is not None and sq[0] not in (None, "")
    has_sq_email = sq is not None and sq[1] not in (None, "")
    check("Config has Site Q row with name in column B",
          has_sq_name,
          f"value={sq!r}" if sq is not None else "row not found")
    check("Config has Site Q row with email in column C",
          has_sq_email,
          f"value={sq!r}" if sq is not None else "row not found")

    # AC 13: LogFile row with non-empty URL in column B
    lf = config_rows.get("LogFile")
    check("Config has LogFile row with non-empty column B",
          lf is not None and lf[0] not in (None, ""),
          f"value={lf!r}" if lf is not None else "row not found")

    st = config_rows.get("Sheet Template")
    check("Config has Sheet Template row with non-empty column B",
          st is not None and st[0] not in (None, ""),
          f"value={st!r}" if st is not None else "row not found")
    if payload is not None and st is not None and st[0] not in (None, ""):
        try:
            config_template_id = extract_sheet_id(str(st[0]).strip())
        except ValueError as err:
            check("Sheet Template row contains a spreadsheet URL", False, str(err))
        else:
            check("Sheet Template row points to template spreadsheet",
                  config_template_id == str(payload.get("templateSpreadsheetId") or "").strip(),
                  f"got={config_template_id!r} expected={payload.get('templateSpreadsheetId')!r}")

    last_month = config_rows.get("Last Month Tracker")
    check("Config has Last Month Tracker row",
          last_month is not None,
          "row not found")
    if template_lineage is not None and last_month is not None:
        expected_previous = template_lineage.get("previous_tracker")
        expected_name = expected_previous["name"] if expected_previous else ""
        expected_url = expected_previous["url"] if expected_previous else ""
        check("Last Month Tracker name matches template Links lookup",
              stringify_cell(last_month[0]) == expected_name,
              f"got={stringify_cell(last_month[0])!r} expected={expected_name!r}")
        check("Last Month Tracker URL matches template Links lookup",
              stringify_cell(last_month[1]) == expected_url,
              f"got={stringify_cell(last_month[1])!r} expected={expected_url!r}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    args = sys.argv[1:]
    log_entry_payload = None

    if len(args) == 1 and args[0] == "--local":
        try:
            sheet_url, log_entry_payload = find_sheet_url_from_local_logs()
        except RuntimeError as e:
            print(f"ERROR: {e}", file=sys.stderr)
            sys.exit(2)
    elif len(args) == 2 and args[0] == "--log":
        log_url = args[1]
        print(f"LogFile URL: {log_url}")
        try:
            sheet_url, log_entry_payload = find_sheet_url_from_log(log_url)
        except RuntimeError as e:
            print(f"ERROR: {e}", file=sys.stderr)
            sys.exit(2)
    elif len(args) == 1 and not args[0].startswith("--"):
        sheet_url = args[0]
    else:
        print(f"Usage:", file=sys.stderr)
        print(f"  {sys.argv[0]} --local", file=sys.stderr)
        print(f"  {sys.argv[0]} --log <drive_logfile_url>", file=sys.stderr)
        print(f"  {sys.argv[0]} <google_sheets_url>", file=sys.stderr)
        sys.exit(2)

    print(f"Sheet URL: {sheet_url}")

    try:
        sheet_id = extract_sheet_id(sheet_url)
    except ValueError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(2)

    export_url = build_export_url(sheet_id)
    print(f"Downloading: {export_url}")

    try:
        xlsx_bytes = download_xlsx(export_url)
    except requests.HTTPError as e:
        print(f"ERROR downloading spreadsheet: {e}", file=sys.stderr)
        sys.exit(2)

    wb = load_workbook(io.BytesIO(xlsx_bytes))
    print(f"Sheets: {wb.sheetnames}")

    if "Tracker" not in wb.sheetnames:
        print("ERROR: 'Tracker' sheet not found", file=sys.stderr)
        sys.exit(2)

    check_tracker(wb["Tracker"])
    check_bonus_tracker(wb)
    check_responses(wb)
    check_activity(wb)
    template_lineage = None
    if log_entry_payload is not None:
        check_log_payload(log_entry_payload)
        template_lineage = check_links_sheet(log_entry_payload)
    check_config(wb, log_entry_payload, template_lineage)

    print(f"\nResults: {RESULT['pass']} passed, {RESULT['fail']} failed")
    sys.exit(0 if RESULT["fail"] == 0 else 1)


if __name__ == "__main__":
    main()
